import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# ── Colours from the template ─────────────────────────────────────
GREEN_HEADER  = '00B050'   # dark green title bar
GREEN_LIGHT   = 'E2EFDA'   # light green alternating rows
GREEN_MID     = '92D050'   # medium green (section headers)
YELLOW_WARN   = 'FFFF00'   # yellow warning bar
BLUE_LINK     = '0070C0'   # email / link colour
WHITE         = 'FFFFFF'
BLACK         = '000000'
DARK_GREEN    = '375623'   # text on green header

def _side(style='thin'):
    return Side(style=style)

def _border(left='thin', right='thin', top='thin', bottom='thin'):
    return Border(
        left=_side(left), right=_side(right),
        top=_side(top),   bottom=_side(bottom)
    )

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _font(bold=False, size=10, color=BLACK, italic=False):
    return Font(name='Arial', size=size, bold=bold,
                color=color, italic=italic)

def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set(ws, row, col, value='', bold=False, size=10, color=BLACK,
         h='left', fill=None, border=None, italic=False, wrap=False,
         num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = _align(h=h, wrap=wrap)
    if fill:   c.fill   = fill
    if border: c.border = border
    if num_fmt: c.number_format = num_fmt
    return c


def generate_invoice_xlsx(invoice) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoice'

    user   = invoice.user
    shifts = list(invoice.shifts.all().order_by('date', 'start_time'))

    MONTHS = ['','January','February','March','April','May','June',
              'July','August','September','October','November','December']
    period_name = MONTHS[invoice.period_month]

    # ── Column widths (A–I) ───────────────────────────────────────
    # A=#  B=DATE  C=CLIENT/EVENT  D=POS  E=TIME_IN  F=TIME_OUT
    # G=HOURS  H=TRAVEL  I=RATE  J=TOTAL
    col_widths = [4, 12, 28, 6, 10, 10, 8, 8, 8, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── ROW 1: Company title + Invoice Period ─────────────────────
    ws.row_dimensions[1].height = 20
    ws.merge_cells('A1:G1')
    _set(ws, 1, 1,
         value='Subcontractors Invoice Form - Staff Your Event Inc.',
         bold=True, size=11, color=WHITE,
         h='left', fill=_fill(GREEN_HEADER),
         border=_border())

    ws.merge_cells('H1:J1')
    c = ws.cell(row=1, column=8)
    c.value        = f'Invoice Period: {period_name}'
    c.font         = _font(bold=True, size=10, color=WHITE)
    c.alignment    = _align(h='right')
    c.fill         = _fill(GREEN_HEADER)
    c.border       = _border()

    # ── ROW 2: "Current Personal Information" ─────────────────────
    ws.row_dimensions[2].height = 16
    ws.merge_cells('A2:J2')
    _set(ws, 2, 1, value='Current Personal Information',
         bold=True, size=10, color=BLACK,
         h='left', fill=_fill(GREEN_MID), border=_border())

    # ── ROW 3: Name + Email ───────────────────────────────────────
    ws.row_dimensions[3].height = 16
    full_name = f'{user.first_name} {user.last_name}'.strip() or user.username

    _set(ws, 3, 1, value='Name:', bold=True, size=9)
    ws.merge_cells('B3:D3')
    _set(ws, 3, 2, value=full_name, size=9,
         fill=_fill(GREEN_LIGHT), border=_border())

    _set(ws, 3, 6, value='Email:', bold=True, size=9)
    ws.merge_cells('G3:J3')
    _set(ws, 3, 7, value=user.email, size=9,
         color=BLUE_LINK, fill=_fill(GREEN_LIGHT), border=_border())

    # ── ROW 4: Address + Home Phone ───────────────────────────────
    ws.row_dimensions[4].height = 16
    address = getattr(user, 'address', '') or ''

    _set(ws, 4, 1, value='Address:', bold=True, size=9)
    ws.merge_cells('B4:D4')
    _set(ws, 4, 2, value=address, size=9,
         fill=_fill(GREEN_LIGHT), border=_border())

    _set(ws, 4, 6, value='Home Phone:', bold=True, size=9)
    ws.merge_cells('G4:J4')
    _set(ws, 4, 7, value='', size=9,
         fill=_fill(GREEN_LIGHT), border=_border())

    # ── ROW 5: City + Cell Phone ──────────────────────────────────
    ws.row_dimensions[5].height = 16
    _set(ws, 5, 1, value='City:', bold=True, size=9)
    ws.merge_cells('B5:D5')
    _set(ws, 5, 2, value='Toronto', size=9,
         fill=_fill(GREEN_LIGHT), border=_border())

    _set(ws, 5, 6, value='Cell Phone:', bold=True, size=9)
    ws.merge_cells('G5:J5')
    _set(ws, 5, 7, value='', size=9,
         fill=_fill(GREEN_LIGHT), border=_border())

    # ── ROW 6: Postal Code + Smart Serve ─────────────────────────
    ws.row_dimensions[6].height = 16
    _set(ws, 6, 1, value='Postal Code:', bold=True, size=9)
    ws.merge_cells('B6:D6')
    _set(ws, 6, 2, value='', size=9,
         fill=_fill(GREEN_LIGHT), border=_border())

    _set(ws, 6, 6, value='Smart Serve Number:', bold=True, size=9)
    ws.merge_cells('G6:J6')
    _set(ws, 6, 7, value='', size=9,
         fill=_fill(GREEN_LIGHT), border=_border())

    # ── ROW 7: Warning bar ────────────────────────────────────────
    ws.row_dimensions[7].height = 16
    ws.merge_cells('A7:J7')
    _set(ws, 7, 1,
         value='INVOICES MUST BE SUBMITTED WITHIN THE FIRST 5 DAYS OF THE NEXT MONTH',
         bold=True, size=10, color=BLACK, h='center',
         fill=_fill(YELLOW_WARN), border=_border())

    # ── ROW 8: Table header ───────────────────────────────────────
    ws.row_dimensions[8].height = 18
    headers = ['#', 'DATE', 'CLIENT/EVENT', 'POS*', 'TIME IN', 'TIME OUT',
               'HOURS', 'TRAVEL', 'RATE', 'TOTAL']
    for col, h in enumerate(headers, 1):
        _set(ws, 8, col, value=h,
             bold=True, size=9, color=WHITE, h='center',
             fill=_fill(GREEN_HEADER), border=_border())

    # ── ROW 9+: Shift data rows ───────────────────────────────────
    DATA_START = 9
    NUM_ROWS   = 34   # template shows 34 data rows

    for i in range(NUM_ROWS):
        r    = DATA_START + i
        ws.row_dimensions[r].height = 15
        row_fill = _fill(GREEN_LIGHT) if i % 2 == 0 else _fill(WHITE)

        if i < len(shifts):
            shift = shifts[i]
            row_vals = [
                i + 1,
                shift.date,
                shift.client or '',
                shift.role or '',
                shift.start_time.strftime('%I:%M %p'),
                shift.end_time.strftime('%I:%M %p'),
                float(shift.hours_worked),
                0,                          # TRAVEL — not in model, default 0
                float(shift.hourly_rate),
                float(shift.total_pay),
            ]
        else:
            row_vals = [i + 1, '', '', '', '', '', '', '', '', '']

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=col, value=val if val != '' or col == 1 else None)
            cell.font      = _font(size=9)
            cell.fill      = row_fill
            cell.border    = _border()
            cell.alignment = _align(h='center' if col in (1,4,5,6,7,8) else 'left')

            if col == 2 and val:
                cell.number_format = 'MMM DD, YYYY'
            if col in (7, 8):
                cell.number_format = '0.00'
            if col == 9:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = _align(h='right')
            if col == 10:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = _align(h='right')
                if i < len(shifts):
                    cell.fill = _fill('CCFFCC')  # highlight total column

    # ── Footer rows ───────────────────────────────────────────────
    LAST_DATA = DATA_START + NUM_ROWS   # row 43

    ws.row_dimensions[LAST_DATA].height = 16

    # "Mark below with X..." + Tax Rate + GST/HST
    ws.merge_cells(f'A{LAST_DATA}:E{LAST_DATA}')
    _set(ws, LAST_DATA, 1,
         value='Mark below with an \'X\' how to receive your cheque',
         size=8, italic=True, border=_border())

    _set(ws, LAST_DATA, 6, value='Tax Rate:', bold=True, size=9,
         h='right', border=_border())
    _set(ws, LAST_DATA, 7,
         value=float(invoice.tax_rate * 100),
         size=9, h='center', fill=_fill(GREEN_LIGHT),
         border=_border(), num_fmt='0.00"%"')

    ws.merge_cells(f'H{LAST_DATA}:I{LAST_DATA}')
    _set(ws, LAST_DATA, 8, value='GST/HST:', bold=True, size=9,
         h='right', border=_border())
    _set(ws, LAST_DATA, 10,
         value=float(invoice.tax_amount),
         size=9, h='right', fill=_fill(GREEN_LIGHT),
         border=_border(), num_fmt='"$"#,##0.00')

    # Total Hours row
    TH_ROW = LAST_DATA + 1
    ws.row_dimensions[TH_ROW].height = 16
    ws.merge_cells(f'A{TH_ROW}:F{TH_ROW}')

    total_hours = sum(float(s.hours_worked) for s in shifts)
    _set(ws, TH_ROW, 7, value='Total Hours:', bold=True, size=9,
         h='right', border=_border())
    _set(ws, TH_ROW, 8, value=round(total_hours, 2),
         bold=True, size=9, h='center',
         fill=_fill(GREEN_LIGHT), border=_border(), num_fmt='0.00')

    # Total $ row
    TOTAL_ROW = LAST_DATA + 2
    ws.row_dimensions[TOTAL_ROW].height = 18
    ws.merge_cells(f'A{TOTAL_ROW}:H{TOTAL_ROW}')
    _set(ws, TOTAL_ROW, 1, value='', fill=_fill(GREEN_LIGHT), border=_border())

    _set(ws, TOTAL_ROW, 9, value='Total $', bold=True, size=11,
         h='right', fill=_fill(GREEN_LIGHT), border=_border())
    _set(ws, TOTAL_ROW, 10, value=float(invoice.total),
         bold=True, size=11, h='right',
         fill=_fill(GREEN_LIGHT), border=_border(),
         num_fmt='"$"#,##0.00')

    # ── POS footnote ──────────────────────────────────────────────
    NOTE_ROW = TOTAL_ROW + 1
    ws.merge_cells(f'A{NOTE_ROW}:J{NOTE_ROW}')
    _set(ws, NOTE_ROW, 1,
         value='*POS = Position worked (W = waiter, S = Supervisor, B = Bartender)',
         size=8, italic=True)

    # ── Freeze panes below header ─────────────────────────────────
    ws.freeze_panes = ws.cell(row=DATA_START, column=1)

    # ── Print area ────────────────────────────────────────────────
    ws.print_area = f'A1:J{NOTE_ROW}'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()